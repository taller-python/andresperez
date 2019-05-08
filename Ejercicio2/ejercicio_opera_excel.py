"""leer el archivo y realizar la respectiva operación de suma entre cada pareja de números"""
import openpyxl

def realizar_opera(hoja, max_fila, max_colu, accion):
    """Se crea funcion que recibe la hoja y la operacion a realizar"""
    # Valor inicial para guardar el resltado
    resultado_opera = ""
    # Definir la operacion
    operaciones = {'SUM':sumar, 'RES':restar, 'MUL':multiplicar, 'DIV':dividir}
    # for para iterar las filas
    for i in range(2, max_fila+1):
        # For para iterar las columnas
        cell = -0
        for j in range(1, max_colu+1):
            # Se guarda el valor en el campo cell
            cell_obj = hoja.cell(row=i, column=j)
            # valida que el tipo sea int
            if isinstance(cell_obj.value, int):
                # Se valida si es la primera columna y solo asigna el valor a Cell
                if cell == -0:
                    cell = cell_obj.value
                else:
                    cell = operaciones[accion](int(cell), int(cell_obj.value))
            else:
                # Si no cumple el tipo int se guarda error y sale del for
                cell = "Error"
                break
        # Se guarda el resultado de la operacion por fila
        resultado_opera = resultado_opera + str(cell) + '\n'
    return resultado_opera

def sumar(num1, num2):
    """ Recibe como Parametros de num1 y num2 y los suma"""
    return num1 + num2

def restar(num1, num2):
    """ Recibe como Parametros de num1 y num2 y los resta"""
    return num1 - num2

def multiplicar(num1, num2):
    """ Recibe como Parametros de num1 y num2 y los multiplica"""
    return num1 * num2

def dividir(num1, num2):
    """ Recibe como Parametros de num1 y num2 y los dividi"""
    if num2 == 0:
        return "Error/0"
    return num1 / num2

#Abrimos el fichero excel
DOCUMENTO = openpyxl.load_workbook("./Ejercicio 2/operaciones.xlsx")

# Se carga la hoja llamada SUMA
HOJA_SUMA = DOCUMENTO.get_sheet_by_name("SUMA")
# Maxima fila
MAX_ROW = HOJA_SUMA.max_row
# Maxima Columna
MAX_COLUMN = HOJA_SUMA.max_column
#se llama a la funcion para realizar la operacion indicada
RESULTADO_SUMA = "SUMA" + '\n' + realizar_opera(HOJA_SUMA, MAX_ROW, MAX_COLUMN, "SUM")

# Se carga la hoja llamada RESTA
HOJA_REST = DOCUMENTO.get_sheet_by_name("RESTA")
# Maxima fila
MAX_ROW = HOJA_REST.max_row
# Maxima Columna
MAX_COLUMN = HOJA_REST.max_column
#se llama a la funcion para realizar la operacion indicada
RESULTADO_REST = "RESTA" + '\n' + realizar_opera(HOJA_REST, MAX_ROW, MAX_COLUMN, "RES")

# Se carga la hoja llamada MULTIPLICACIÓN
HOJA_MULT = DOCUMENTO.get_sheet_by_name("MULTIPLICACIÓN")
# Maxima fila
MAX_ROW = HOJA_MULT.max_row
# Maxima Columna
MAX_COLUMN = HOJA_MULT.max_column
#se llama a la funcion para realizar la operacion indicada
RESULTADO_MULT = "MULTIPLICACIÓN" + '\n' + realizar_opera(HOJA_MULT, MAX_ROW, MAX_COLUMN, "MUL")

# Se carga la hoja llamada DIVISIÓN
HOJA_DIVI = DOCUMENTO.get_sheet_by_name("DIVISIÓN")
# Maxima fila
MAX_ROW = HOJA_DIVI.max_row
# Maxima Columna
MAX_COLUMN = HOJA_DIVI.max_column
#se llama a la funcion para realizar la operacion indicada
RESULTADO_DIVI = "DIVISIÓN" + '\n' +  realizar_opera(HOJA_DIVI, MAX_ROW, MAX_COLUMN, "DIV")

# Se guarda todo el resultado en una sola variable
RESULT = RESULTADO_SUMA + RESULTADO_REST + RESULTADO_MULT + RESULTADO_DIVI

# Se crea un archivo .txt
FILE = open("./Resultado.txt", "w")
# Se le agrega el resultado
FILE.write(RESULT)
# Se cierra la edicion del archivo
FILE.close()
print(RESULT)
